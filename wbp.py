import requests
import os
import json
import csv
from datetime import datetime
import tkinter as tk
from tkinter import *
import xlsxwriter 
from openpyxl import load_workbook


def find_id(article):

    if len(article) == 8 :
        article = str(article)
        article3 = article[0:3] 
        article5 = article[0:5]
    if len(article) == 9 :
        article = str(article)
        article3 = article[0:4] 
        article5 = article[0:6]

    url1 = "https://basket-01.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url2 = "https://basket-02.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url3 = "https://basket-03.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url4 = "https://basket-04.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url5 = "https://basket-05.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url6 = "https://basket-06.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url7 = "https://basket-07.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url8 = "https://basket-08.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url9 = "https://basket-09.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"
    url10 = "https://basket-10.wb.ru/vol" + article3 + "/part" + article5 + "/" + str(article) + "/info/ru/card.json"

    response = requests.get(url=url1)
    if response.status_code == 404 :
        response = requests.get(url=url2)
    if response.status_code == 404 : 
        response = requests.get(url=url3)
    if response.status_code == 404 : 
        response = requests.get(url=url4)
    if response.status_code == 404 : 
        response = requests.get(url=url5)
    if response.status_code == 404 : 
        response = requests.get(url=url6)
    if response.status_code == 404 :
        response = requests.get(url=url7)
    if response.status_code == 404 : 
        response = requests.get(url=url8)
    if response.status_code == 404 : 
        response = requests.get(url=url9)
    if response.status_code == 404 : 
        response = requests.get(url=url10)

    try:
        with open(f'info_{article}.json', 'w') as file:
            json.dump(response.json(), file, indent=4, ensure_ascii=False)

        curr_id = response.json()['imt_id']
    
        file = "info_" + str(article) + ".json"
        file_dest(file)
        return curr_id
    except:
        print("id ne naiden")
        file = "info_" + str(article) + ".json"
        file_dest(file)
        return 0

def collect_raiting(article, id):
    t_date = datetime.now().strftime("%d_%m_%Y")
    url1 = "https://feedbacks1.wb.ru/feedbacks/v1/" + str(id)
    url2 = "https://feedbacks2.wb.ru/feedbacks/v1/" + str(id)
    response = requests.get(url=url1)
    if response.json()["feedbackCount"] == 0 :
        response = requests.get(url=url2)
    try:
        with open(f'info_{id}.json', 'w') as file:
            json.dump(response.json(), file, indent=4, ensure_ascii=False)
    except:
        print("plohoi otvet po id")

    result = []

    try:
        valuation = response.json()['valuation']
        feedbackCount = response.json()["feedbackCount"]
        total_valuations = response.json()['valuationDistribution']
        val1 = total_valuations['1']
        val2 = total_valuations['2']
        val3 = total_valuations['3']
        val4 = total_valuations['4']
        val5 = total_valuations['5']
    except:
        print("oshibka obrabotki")
        valuation = "X" 
        feedbackCount = "X"
        total_valuations = 'X'
        val1 = 'X'
        val2 = 'X'
        val3 = 'X'
        val4 = 'X'
        val5 = 'X'
    result.append(article)
    result.append(valuation)
    result.append(feedbackCount)
    result.append(val1)
    result.append(val2)
    result.append(val3)
    result.append(val4)
    result.append(val5)

    append_data(result)
    file = "info_" + str(id) + ".json"
    file_dest(file)

def file_dest(filename): 
    if os.path.exists(filename):
        os.remove(filename)
    else :
        print("net faila")

def catch_articles(spisok):

    stack = []
    index1 = 0
    index2 = 0
    while index2 != len(spisok)-1 and index2 != len(spisok) and index2 >=0 :
        index2 = spisok.find("\n",index1)
        stack.append(spisok[index1:index2])
        index1=index2+1

    return stack

def create_result_file():
    workbook = xlsxwriter.Workbook('result.xlsx')
    worksheet = workbook.add_worksheet('result')
    workbook.close()

def create_header():
    wb = load_workbook('result.xlsx')
    ws = wb['result']
    ws.append(['Артикул','Общ.рейт','Общ.колво','1','2','3','4','5'])
    wb.save('result.xlsx')
    wb.close()  

def append_data(result):
    wb = load_workbook('result.xlsx')
    ws = wb['result']  
    ws.append(result)  
    wb.save('result.xlsx')   
    wb.close() 


def window():

    win = tk.Tk()
    win.title("Parse WB easily")
    win.geometry("700x500+200+100")

    label1 = tk.Label(win, text = "Введите артикулы - каждый с новой строки", 
                      bg ="blue", 
                      relief =tk.RAISED)
    label1.pack()

    field = Text()
    field.pack()

    label2 = tk.Label(win, text = 'Нажмите "ЗАПУСК" и подождите в 2 раза меньше секунд, чем количество введенных артикулов', 
                      bg ="blue", 
                      relief =tk.RAISED)
    label2.pack()

    label3 = tk.Label(win, text = 'Или пока кнопка "ЗАПУСК" не потухнет ;-)', 
                      bg ="blue", 
                      relief =tk.RAISED)
    label3.pack()
 
    def b1_click():
        button1.config(state =tk.DISABLED)
        file_dest('result.xlsx')
        create_result_file()
        create_header()      

        theText = field.get('1.0', "end")
        abba = catch_articles(theText)
        for item in abba :
            id = find_id(item)
            collect_raiting(item, id)

    button1 = tk.Button(win, text = "ЗАПУСК",
                    bg ="red",
                    activebackground = "red",
                    command =b1_click)
    button1.pack()

    win.mainloop()


def main():
    window()
    
if __name__ == '__main__':
    main()


